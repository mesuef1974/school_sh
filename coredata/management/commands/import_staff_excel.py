import os
from django.core.management.base import BaseCommand
from django.db import transaction
from coredata.models import Staff, JobTitle

# سنعتمد openpyxl لقراءة ملفات xlsx
try:
    from openpyxl import load_workbook
except Exception as e:
    load_workbook = None


NORMALIZE_MAP = {
    "مدرس": "معلم",
    "امين": "أمين",
    "امين مخزن": "أمين مخزن",
    "مشرف اداري": "مشرف إداري",
}

def normalize_ar_text(text: str) -> str:
    if not text:
        return text
    t = str(text).strip()
    # توحيد همزات ومسافات بسيطة
    t = t.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    t = " ".join(t.split())
    return t


def normalize_job_title(title: str) -> str:
    if not title:
        return title
    raw = str(title).strip()
    base = normalize_ar_text(raw)
    # تطبيق المرادفات
    for k, v in NORMALIZE_MAP.items():
        if k in base:
            base = base.replace(k, v)
    # استعادة بعض الهمزات الشائعة بعد التطبيع العام
    base = base.replace("امين", "أمين").replace("اداري", "إداري")
    return base


def key_or_none(d, keys):
    for k in keys:
        if k in d and d[k] not in (None, ""):
            return d[k]
    return None


class Command(BaseCommand):
    help = "Import/Update Staff data from an Excel file (xlsx) and fill missing fields only by default."

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, default=r'D:\SULTAN_HAGRY\shaniya_django\Doc_School\school_DATA\stuff_03.xlsx', help='Path to Excel file (.xlsx)')
        parser.add_argument('--sheet', type=str, default=None, help='Worksheet name. If omitted, uses the first sheet')
        parser.add_argument('--overwrite', action='store_true', help='Overwrite existing non-empty Staff fields with Excel values')

    def handle(self, *args, **options):
        if load_workbook is None:
            self.stdout.write(self.style.ERROR('openpyxl غير مثبت. رجاءً ثبت openpyxl ثم أعد المحاولة.'))
            return

        file_path = options['file']
        sheet_name = options['sheet']
        overwrite = options['overwrite']

        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f"لم يتم العثور على الملف: {file_path}"))
            return

        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            self.stdout.write(self.style.ERROR("الملف خالٍ من البيانات."))
            return

        # بناء خريطة رؤوس مرنة
        header_map = {}
        for idx, h in enumerate(headers):
            key = str(h or '').strip()
            key_norm = key.lower().replace(' ', '').replace('_', '')
            header_map[idx] = key_norm

        def row_to_dict(vals):
            d = {}
            for i, v in enumerate(vals):
                d[header_map.get(i, f'col{i}')]= v
            return d

        # مفاتيح الأعمدة المتوقعة بأشكال مختلفة
        KEY_NAME = ['name', 'stuffname', 'stuff_name', 'studentname', 'الاسم']
        KEY_JOB_TITLE = ['jobtitle', 'title', 'job_title', 'المسمىالوظيفي']
        KEY_JOB_NO = ['jobno', 'job_number', 'jobnumber', 'job_no', 'job', 'jobid']
        KEY_EMAIL = ['email', 'mail', 'البريدالالكتروني', 'الايميل']
        KEY_NATIONAL = ['nationalno', 'national_id', 'nationalid', 'nationalnumber', 'national_no']
        KEY_PHONE = ['phoneno', 'phone', 'mobile', 'mobileno', 'phone_no']
        KEY_ACCOUNT_NO = ['accountno', 'account_no', 'bankaccount', 'account', 'رقمالحساب']
        KEY_USERNAME = ['username', 'user']

        created = 0
        updated = 0
        skipped = 0

        with transaction.atomic():
            for vals in rows:
                if not vals:
                    continue
                rd = row_to_dict(vals)

                name = key_or_none(rd, KEY_NAME)
                job_title = key_or_none(rd, KEY_JOB_TITLE)
                job_no = key_or_none(rd, KEY_JOB_NO)
                email = key_or_none(rd, KEY_EMAIL)
                national_no = key_or_none(rd, KEY_NATIONAL)
                phone_no = key_or_none(rd, KEY_PHONE)
                account_no = key_or_none(rd, KEY_ACCOUNT_NO)
                username = key_or_none(rd, KEY_USERNAME)

                # تخطي الصفوف غير الصالحة
                if not name and not (job_no or email or national_no):
                    skipped += 1
                    continue

                name = str(name).strip() if name else None
                job_no = str(job_no).strip() if job_no else None
                email = str(email).strip() if email else None
                national_no = str(national_no).strip() if national_no else None
                phone_no = str(phone_no).strip() if phone_no else None
                account_no = str(account_no).strip() if account_no else None
                username = str(username).strip() if username else None

                # تطبيع المسمى
                job_title_norm = normalize_job_title(job_title) if job_title else None

                # محاولة مطابقة سجل قائم
                staff_qs = Staff.objects.all()
                staff_obj = None
                for key, value in (("national_no", national_no), ("job_number", job_no), ("email", email)):
                    if value:
                        try:
                            staff_obj = staff_qs.get(**{key: value})
                            break
                        except Staff.DoesNotExist:
                            pass
                        except Staff.MultipleObjectsReturned:
                            # في حال التكرار الشاذ، نعتمد أول واحد
                            staff_obj = staff_qs.filter(**{key: value}).first()
                            break

                if staff_obj is None and name:
                    try:
                        staff_obj = staff_qs.get(name=name)
                    except Staff.DoesNotExist:
                        staff_obj = None

                # تجهيز JobTitle
                job_title_obj = None
                if job_title_norm:
                    job_title_obj, _ = JobTitle.objects.get_or_create(title=job_title_norm)

                # إنشاء أو تحديث
                if staff_obj is None:
                    staff_obj = Staff(
                        name=name or (email or job_no or national_no or "")
                    )
                    # تعبئة الحقول المتاحة
                    staff_obj.job_title = job_title_obj if job_title_obj else None
                    staff_obj.job_number = job_no or None
                    staff_obj.email = email or None
                    staff_obj.national_no = national_no or None
                    staff_obj.phone_no = phone_no or None
                    staff_obj.account_no = account_no or None
                    staff_obj.save()
                    created += 1
                else:
                    # تحديث انتقائي أو شامل حسب الخيار
                    def set_field(obj, field, value):
                        if value in (None, ""):
                            return False
                        current = getattr(obj, field)
                        if overwrite or (current in (None, "")):
                            setattr(obj, field, value)
                            return True
                        return False

                    changed = False
                    changed |= set_field(staff_obj, 'name', name)
                    if job_title_obj:
                        if overwrite or (staff_obj.job_title is None):
                            staff_obj.job_title = job_title_obj
                            changed = True
                    changed |= set_field(staff_obj, 'job_number', job_no)
                    changed |= set_field(staff_obj, 'email', email)
                    changed |= set_field(staff_obj, 'national_no', national_no)
                    changed |= set_field(staff_obj, 'phone_no', phone_no)
                    changed |= set_field(staff_obj, 'account_no', account_no)

                    if changed:
                        staff_obj.save()
                        updated += 1
                    else:
                        skipped += 1

        self.stdout.write(self.style.SUCCESS(f"تم التنفيذ. مضاف: {created}, محدّث: {updated}, متخطّى: {skipped}"))